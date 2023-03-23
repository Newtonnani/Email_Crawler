import time, json, sys, os, boto3, extract_msg, requests, traceback, pandas as pd, re, urllib3, platform, shutil, gzip, \
    uuid
from requests.auth import HTTPBasicAuth
from datetime import datetime, timedelta
from collections import Counter
import urllib.parse
import xlrd, openpyxl
import pandas as pd
import numpy as np
from datetime import date
import pgSQL
from openpyxl import load_workbook

sql_w = pgSQL.SQL(random_host=False)
sql_r = pgSQL.SQL(random_host=True)

path_separator = '/' if 'linux' in str(platform.system()).lower() else '\\'
tmp_downloads = None
step_log = {}
step_log_status = []
df = None
stop_process = False
error = None
target_month = None
filename = ""
tmp_downloads = '/tmp'

s3_client = boto3.client('s3')


def handler(event, context):
    try:
        main_fn(event)
    except Exception:
        print(str(traceback.format_exc().splitlines()))


def main_fn(event):
    global bucket, key, job_id, email_id, received_date, source, source_id, tmp_downloads, cfg, f_path, contact, filename
    source = 'scraper'
    source_id = str(uuid.uuid4())
    try:

        obj = event["Records"][0]
        # parse object to get bucket and key
        bucket = str(obj['s3']['bucket']['name'])
        key = urllib.parse.unquote_plus(str(obj["s3"]['object']['key']))
        s3 = boto3.resource('s3')
        src_bucket = s3.Bucket(bucket)

        # check if folder creation event
        if key.endswith('/'):
            print('Folder creation, ignored!')
        else:
            keys = key.split('/')
            job_id = keys[1]
            email_id = keys[2]  # (key.split('/')[-1]).split('.')[0]
            if len(keys) > 4:
                received_date = keys[3]
            filename = keys[-1]
            print(keys)
            # get db config
            jdoc = {}
            jdoc['job_id'] = job_id
            jdoc['email_id'] = email_id

            res = db_query(jdoc, 'config')
            print(res)

            if res["response"]["status"] == "ok":
                # source_id = res["scraper_id"]
                cfg = res["worker_config"]
                contact = res["it_contact"]
                step_log['job_id'] = job_id
                step_log['email_id'] = email_id
                step_log['email_details'] = res["email_details"]
                # print(cfg)

                # update db scraper_exec_status
                rs = {}
                rs = {"update_task": 'scraper_exec_status', "job_id": job_id, "source_id": source_id, "data": "Running"}
                db_query(rs)
                # update logs
                rs = {}
                rs = {"update_task": 'update_log', "job_id": job_id, "source_id": source_id, "source": source,
                      "log_type": "INFO", "log_details": [{"status": "ok", "msg": "config dispatched"}]}
                db_query(rs)

                # tmp_downloads = local_path + source_id
                # if not os.path.exists(tmp_downloads):
                #    os.makedirs(tmp_downloads)
                # download file from s3
                f_path = tmp_downloads + path_separator + filename
                step_log['file_path'] = f_path
                src_bucket.download_file(key, f_path)
                print("s3 file downloaded")
                if filename.endswith(".gz"):
                    unzip(f_path, f_path.replace('.gz', ''))
                os.remove(tmp_downloads + path_separator + filename)
                # loop thru processing steps
                step_loop_thru(cfg)

                print(step_log)
                print(step_log_status)
                data = 'COMPLETED'
                logtype = 'INFO'
                for s in step_log_status:
                    if s['status'] != 'OK':
                        data = 'ERROR'
                        logtype = 'ERROR'
                # update logs
                rs = {}
                rs = {"update_task": 'update_log', "job_id": job_id, "source_id": source_id, "source": source,
                      "log_type": logtype, "log_details": [{"status": "ok", "msg": step_log_status}]}
                db_query(rs)
                if logtype == 'ERROR':
                    send_email()

        # update db scraper_exec_status
        rs = {}
        rs = {"update_task": 'scraper_exec_status', "job_id": job_id, "data": data}
        db_query(rs)

    except Exception:
        print(str(traceback.format_exc().splitlines()))
        print(step_log)
        print(step_log_status)
        # update logs
        rs = {}
        rs = {"update_task": 'update_log', "job_id": job_id, "source_id": source_id, "source": source,
              "log_type": "ERROR", "log_details": [{"status": "error", "msg": step_log_status}]}
        db_query(rs)
        # update db scraper_exec_status
        rs = {}
        rs = {"update_task": 'scraper_exec_status', "job_id": job_id, "email_id": email_id, "data": "ERROR"}
        db_query(rs)
        # send mail
        send_email()


def unzip_zipping_file(doc):
    # zipped file in email extracts and delete the file too
    global tmp_downloads
    from zipfile import ZipFile
    try:
        if doc["process_zipped_file"] and doc["process_zipped_file"] is not None:
            file_path = step_log["file_path"]
            with ZipFile(file_path.replace(".gz", ""), 'a') as zipobj:
                zipobj.extractall(tmp_downloads)
            os.remove(file_path.replace(".gz", ""))
            status_update(doc)
            return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def unzip(filename, unzipped_file):
    try:
        with gzip.open(filename, 'rb') as f_in:
            with open(unzipped_file, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)
    except Exception:
        print(traceback.format_exc().splitlines())


def file_push_to_s3(doc):
    try:
        dt_val = str(date.today()) if target_month is None else str(target_month)
        s3_client = boto3.client('s3')
        sub_path = ''
        dt = dt_val.split('-')
        if 'subpath' in doc.keys() and doc['subpath'] is not None:
            for st in doc['subpath']:
                try:
                    f_val = str(step_log[st]).lower()  # str(re.sub(r'\W', '_', str(step_log[st]).lower()))
                    sub_path += (f_val + '/') if len(f_val) > 0 else ''
                except Exception:
                    pass

        sub_path = sub_path.replace('#yyyy#', dt[0]).replace('#mm#', dt[1]).replace('#dd#', dt[2])
        sub_path = sub_path.replace('#jobid#', job_id).replace('#emailid#', email_id)

        s3_client.upload_file(step_log[doc['link_to_step']], doc['bucket'],
                              doc['path'] + sub_path + (step_log[doc['link_to_step']].split(path_separator)[-1]))
        status_update(doc)

        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def db_query(rs, typ=None):
    try:
        if typ == 'config':
            q = "select email_crawler.worker_get_config('{}')".format(json.dumps(rs))
        else:
            q = "select email_crawler.do_updates('{}')".format(str(json.dumps(rs)).replace("'", "''"))
        # print(q)
        sql_w.conn()
        rs = sql_w.executeAndReturn(q)
        sql_w.close_conn()
        return rs[0][0]
    except Exception:
        doc = {}
        doc['step_id'] = 'DB-ERROR'
        status_update(doc, traceback.format_exc().splitlines())
        print(str(traceback.format_exc().splitlines()))
        # update logs
        rs = {}
        rs = {"update_task": 'update_log', "job_id": job_id, "source_id": source_id, "source": source,
              "log_type": "ERROR",
              "log_details": [{"status": "error", "msg": str(traceback.format_exc().splitlines())}]}
        db_query(rs)
        # update status
        rs = {}
        rs = {"update_task": 'crawler_exec_status', "job_id": job_id, "data": "Error"}
        db_query(rs)


def file_get_downloaded(doc):
    # this function will get the full path of a downloaded file.
    global tmp_downloads
    tmp = None
    timeout = 150
    counter = 0
    while True:
        tmp = os.listdir(tmp_downloads)
        if len(tmp) > 0:  # and tmp[0].split('.')[-1] != 'crdownload':
            step_log[doc['step_id']] = tmp_downloads + path_separator + tmp[0]
            status_update(doc)
            return True
        else:
            time.sleep(2)
            if counter >= timeout:
                status_update(doc, traceback.format_exc().splitlines())
                return False
            counter = counter + 1


def dataframe_insert_file_name(doc):
    global df, filename
    try:
        index = doc['column']['index'] if doc.get('column').get('index') else 0
        column_name = doc['column']['name'] if doc.get('column').get('name') else "input_file_name"
        val = filename
        df.insert(
            loc=index,
            column=column_name,
            value=val
        )
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_sams_club_revenue(doc):
    global df
    try:
        f_path = step_log[doc['link_to_step']]
        data = pd.ExcelFile(f_path, engine='openpyxl')
        dfs = data.parse(doc['file_details']['sheet_name'])

        #############################
        # TO GET THE DATE IN OTHER SHEET
        #############################

        dfs_to_date = data.parse(doc['file_details']['extract_date_sheet_name'])
        mask = [value for value, key in enumerate(dfs_to_date.columns) if 'Fis' in key]
        first_row_file = list(dfs_to_date.columns)[mask[0]]
        curr_date = first_row_file.split(':')[1]
        start_date = re.search(r"\d{1,2}-\d{1,2}-\d{2,4}", str(curr_date)).group()
        seven_days_date = pd.DataFrame(pd.date_range(start=start_date, periods=7).strftime('%m/%d/%Y'),
                                       columns=['Date'])

        #############################
        # CHECK FOR "Category Number" IN THE SHEET
        #############################

        column_index = 0

        # getting columns names and index of this column

        for idx, d in enumerate(dfs.to_dict('records')):
            if str(d.get("Unnamed: 0")).strip() == 'Category Number':
                column_index = idx
        dataframe = data.parse(doc['file_details']['sheet_name'], skiprows=column_index + 1, header=0)

        dataframe = dataframe[dataframe["Item Number"].notna()]

        #############################
        # COLUMNS FORMAT
        #############################
        name_columns = dataframe.columns
        new_name_columns = name_columns.str.replace('%', 'PCT').str.replace('$', 'DLL').str.replace('\n',
                                                                                                    ' ').str.replace(
            '.', '').str.replace(' ', '_').str.replace('_&_', '_')
        new_name_columns = new_name_columns.str.replace(' ', '_')
        dataframe.columns = new_name_columns

        #############################
        # DIVIDE INFORMATION
        #############################

        columns_to_normalize = ['Sales_Unit_Qty', 'Sales_Unit_Qty_YA', 'Sales_DLL', 'Sales_DLL_YA',
                                'B&M_Sales_Unit_Qty',
                                'B&M_Sales_Unit_Qty_YA', 'B&M_Sales_DLL', 'B&M_Sales_DLL_YA',
                                'Ecomm_Completed_Sales_Unit_Qty', 'Ecomm_Completed_Sales_Unit_Qty_YA',
                                'Ecomm_Completed_Sales_DLL', 'Ecomm_Completed_Sales_DLL_YA']

        # columns with not percentages
        mask_int_or_float = [idx for idx, value in enumerate(dataframe.columns) if value in columns_to_normalize]

        # columns to normalize
        arr_columns = list(dataframe.columns)  # name of columns to normalize
        raw_iteration = list()

        # Normalization raw by raw
        for raw in range(dataframe.shape[0]):
            arr = dataframe.iloc[raw, :].copy()  # select raw

            operation = np.array(arr.iloc[mask_int_or_float] / 7)  # divide columns of single raw by 7

            arr.iloc[mask_int_or_float] = operation  # change values

            # repeat values seven times
            arr_repeat = pd.DataFrame(list(map(lambda x: arr.values.tolist(), range(7))), columns=arr_columns)

            arr_repeat['Day'] = seven_days_date  # insert column with seven days

            mask_order_columns = [arr_repeat.columns[-1]] + list(arr_repeat.columns)[:-1]  # index to sort columns

            arr_repeat = arr_repeat[mask_order_columns]  # sort columns

            raw_iteration.append(arr_repeat)  # append blocks

        df_result = pd.concat(raw_iteration).reset_index().iloc[:, 1:]
        df = df_result
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_staples_segment(doc):
    global df, target_month
    try:
        f_path = step_log[doc['link_to_step']]
        data = pd.ExcelFile(f_path, engine='openpyxl')
        xl_tabs_ = doc['file_details']['xl_tabs']
        col_names = doc['file_details']['col_names']
        section_ = doc['file_details']['section']
        sub_section_ = doc['file_details']['sub_section']

        vals_ = list(col_names.keys())  # ["HP Sales $","Staples Sales $"]#,"HP Units",'Staples Units']
        recs = []
        # target_month = str(date.today().replace(day=1).replace(month=int(date.today().month)-1 if int(date.today().month) != 1 else 12))
        target_month = str(date.today().replace(day=1).replace(
            month=int(date.today().month) - 1 if int(date.today().month) != 1 else 12).replace(
            year=int(date.today().year) - 1 if int(date.today().month) == 1 else int(date.today().year)))
        # for each table
        for tb in xl_tabs_:
            # var to store the col name of the target col
            cur_col = None

            # load xlsx tab
            dfs = data.parse(tb)

            # get 1st row containing the months name
            dd = [d for d in dfs.to_dict('records') if d[dfs.columns[0]] in ['Total', 'Total Toner & Ink']][0]

            # find the column for the current month (target col)
            for k in dd.keys():
                if str(dd[k]).split()[0] == target_month:
                    cur_col = k

            # Keep only 1st col and the target col and rename them to manageble col names
            col_map = {dfs.columns[0]: 'keycol', cur_col: 'valcol'}
            df = dfs[[dfs.columns[0], cur_col]].rename(columns=col_map)

            # remove unused df to save memory
            del dfs

            # Keep only rows which on 1st coll contain valuse of interest
            dt = df[df.keycol.isin(section_ + sub_section_ + vals_)].to_dict('records')

            # Build individual records
            # for each record in the list
            for i in range(len(dt)):
                # if the value is part of the section list
                if dt[i]['keycol'] in section_:
                    # for a predicted number of rows check if the value is part of the sub section list
                    for j in range(
                            i,
                            min(
                                i + 1 + ((len(sub_section_) * len(vals_)) * 2),
                                len(dt)
                            )
                    ):
                        # if the value is part of the sub section list create a record
                        if dt[j]['keycol'] in sub_section_:
                            nr = {}

                            nr[
                                'category'] = 'CPS' if 'Computing' in tb else 'Print HW' if 'Printing' in tb else 'Supplies' if 'Toner' in tb else ''
                            nr['sub_category'] = 'Toner' if 'toner' in dt[j]['keycol'].lower() \
                                else 'Ink' if 'ink' in dt[j]['keycol'].lower() \
                                else 'Ink+Toner' if 'combined' in dt[j]['keycol'].lower() \
                                else 'Inkjet' if 'inkjet printing' in dt[i]['keycol'].lower() \
                                else 'Laserjet' if 'laser printing' in dt[i]['keycol'].lower() \
                                else 'Notebook' if 'notebook' in dt[i]['keycol'].lower() \
                                else 'Desktop' if 'desktop' in dt[i]['keycol'].lower() \
                                else ''

                            nr['xl_tab'] = tb
                            nr['month'] = target_month
                            nr['section'] = dt[i]['keycol']
                            nr['sub_section'] = dt[j]['keycol']
                            # for a known number of values check if the following records contain the expected values
                            for k in range(j, min(j + len(vals_) + 1, len(dt))):
                                # if col name in the vals_ list pull it's value
                                if dt[k]['keycol'] in vals_:
                                    nr[col_names[dt[k]['keycol']]] = dt[k]['valcol']
                            recs.append(nr)

                            # remove unused dict to save memory
            del dt

        df = pd.DataFrame(recs)
        step_log[doc['step_id']] = str(list(col_names.keys())[0]).split(' ', 1)[-1].split(' ')[0]
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_ink_toner_online_metrics(doc):
    global df
    try:
        data = pd.ExcelFile(step_log[doc['link_to_step']], engine='openpyxl')
        df = data.parse(step_log[doc["file_details"]["xl_tab_from_step"]], skiprows=[0], nrows=12, header=None)
        recs = df.to_dict('records')
        for r in recs:
            cur_col = r[0]
            if cur_col.lower() == 'ly':
                r[0] = prev_col + ' ' + cur_col
            prev_col = r[0]

        df = pd.DataFrame(recs).set_index(0)
        df = df.transpose()

        for col in doc["file_details"]["filter_na_on_cols"]:
            df = df[df[col].notna()]

        df.rename(columns=doc["file_details"]['col_names'], inplace=True)

        for col in doc["file_details"]['re_format_date_col']['col_names']:
            df[col] = pd.to_datetime(df[col], infer_datetime_format=True).dt.strftime(
                doc["file_details"]['re_format_date_col']['format'])
        df = df.fillna('0')
        for column in df:
            if df[column].name != 'week_start' and df[column].name != 'week_end' and df[column].name != 'ink' and df[
                column].name != 'toner':
                df[column] = pd.to_numeric(df[column].replace('na', '0').astype('Int64', errors='ignore'))

        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_best_buy_monthly_data(doc):
    global df

    try:
        file_type = doc["file_type"]
        if file_type == "RMP":
            mapping = {
                'Sponsored Rmp Summary': 'Date',
                'By Product': 'SKU',
                'By Keyword': 'Keyword',
                'By Placement': 'Placement Type',
                'By Platform': 'Platform',
                'By Opportunities': 'Campaign Name'

            }
        elif file_type == "Paid_Search":
            mapping = {
                'Paid Search Summary': 'Date',
                'By Campaign': 'Campaign Type',
                'By Query': 'Search Query',
                'By Device': 'Device',
                'By Product': 'SKU'
            }

        combined_df = pd.DataFrame()
        df_collection = []
        sheet = step_log[doc["file_details"]["sheet_name"]]
        column = mapping[sheet]
        file_path = step_log[doc['link_to_step']]
        wb = load_workbook(file_path)
        wb_sheet = wb[sheet]
        column_name = wb_sheet["B5"].value
        # column_value = wb_sheet["c5"].value
        report_date = wb_sheet["c11"].value
        if not report_date:
            report_date = str(wb_sheet["c12"].value)
        if (column_name is not None) or (column_name == "Campaign"):
            column_value = str(wb_sheet["c5"].value)
        else:
            column_name = str(wb_sheet["B6"].value)
            column_value = str(wb_sheet["c6"].value)

        # if file_type == "RMP" and sheet != "Sponsored Summary":
        #     column_name = wb_sheet["B5"].value
        #     column_value = wb_sheet["c5"].value

        local_df = pd.read_excel(file_path, sheet_name=sheet, engine='openpyxl')
        start_table_index = local_df.index[local_df['Unnamed: 1'] == column].tolist()[0]
        local_df = pd.read_excel(file_path, sheet_name=sheet, skiprows=range(2, start_table_index), engine='openpyxl')
        local_df.drop('Unnamed: 0', axis=1, inplace=True)
        local_df.columns = local_df.iloc[2]
        sheet_columns = list(local_df.columns)

        local_df.drop([0, 2], axis=0, inplace=True)

        first_row_df = local_df.iloc[0]
        if file_type == "Paid_Search":
            format_columns = ['Online', 'In Store', 'Total']
            parent_columns = ['Revenue', 'ROAS']

        else:
            format_columns = ['Units', 'Revenue', 'ROAS ($)', 'ROAS (%)',
                              'Capout Time', 'Number of Active SKUs', '% of Traffic Active',
                              'Missed Clicks', 'Click Opportunity', 'Delivered Spend', 'Missed Spend',
                              'Spend Opportunity', 'Missed Revenue', 'Revenue Opportunity'
                              ]
            parent_columns = ['Direct Click', 'Direct View', 'Related Click', 'Total', 'Click', 'View Through',
                              'Click + View', 'Daily Averages']
        if sheet == 'By Opportunities':
            format_columns.append('Clicks')
        for parent_column in parent_columns:
            for columns in format_columns:
                if columns in local_df.columns:
                    if parent_column in first_row_df.values and columns in sheet_columns:
                        sheet_columns[sheet_columns.index(columns)] = '{0} - {1}'.format(parent_column, columns)

        local_df.drop([1], axis=0, inplace=True)
        local_df.columns = sheet_columns
        local_df[column_name] = column_value
        local_df["input_file_name"] = file_path.split("/")[-1]
        local_df = local_df[:-1]
        local_df = local_df.replace(['-'], 0)
        if not re.search('[a-zA-Z]', report_date) and (
                not (sheet == 'Paid Search Summary' or sheet == 'Sponsored Rmp Summary')):
            local_df.insert(loc=0, column='Date', value=report_date.split()[0])

        if not local_df.empty:
            df_collection.append(local_df)

        # print("*" * 100)
        # print(sheet)
        # print(df_collection[0])
        # print(df_collection[0].shape)
        if df_collection:
            df = df_collection[0]
        else:
            df = pd.DataFrame(columns=sheet_columns)
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def cal_date_delta(doc):
    'example : { "delta": "delta-", "step_id": "getting_the_required_date","step_action": "cal_date_delta","received_date": null,"requested_day_name": "sat"}'
    try:
        from calendar import day_name
        weekend_name = {"mon": 0, "tue": 1, "wed": 2, "thur": 3, "fri": 4, "sat": 5, "sun": 6}
        week_name = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

        if doc["received_date"] and doc["received_date"] is not None:
            given_received_date = datetime.fromisoformat(doc["received_date"])
            if doc["delta"] and doc["delta"] is not None:
                prev = doc["delta"].split("-")
                next = doc["delta"].split("+")
                if len(prev) == 2:
                    requested_date = given_received_date - timedelta(days=prev[-1])
                if len(next) == 2:
                    requested_date = given_received_date + timedelta(days=next[-1])
        else:
            requested_day_name = day_name[weekend_name[doc["requested_day_name"]]]
            delta = doc["delta"]
            email_date = datetime.fromisoformat(received_date)
            email_day_name = day_name[email_date.weekday()]
            if delta == 'delta-':
                week_name = week_name[::-1]
            else:
                week_name = week_name
            index = [i for i in range(len(week_name)) if week_name[i] == email_day_name][0]
            # calculating days for that requested_day_name
            days_to_reduced = 0
            while email_day_name != requested_day_name:
                index = (index + 1) % len(week_name)
                email_day_name = week_name[index]
                days_to_reduced += 1

            if delta == 'delta-':
                requested_date = email_date - timedelta(days=days_to_reduced)
            if delta == 'delta+':
                requested_date = email_date + timedelta(days=days_to_reduced)

        final_date = requested_date.date()
        step_log[doc['step_id']] = final_date
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_staples_attach_reporting(doc):
    # developed for ATE-31
    # after unzippped file later processing with excel file
    """ Eg:
    {
		"step_id": "dataframe_load",
		"step_action": "dataframe_load_staples_attach_reporting",
		"file_details": {
			"type": "xlsx",
			"sheet_names": ["Ink","Toner"],
			"columns": ["week_of", "attach_by_vendor", "total", "hewlett_packard"]
		},
		"link_to_step": "loading_file"
	}
    """
    global df
    import openpyxl
    try:
        list_df = []
        book = openpyxl.load_workbook(filename=step_log[doc['link_to_step']], read_only=True,
                                      data_only=True)
        # df_obj = pd.ExcelFile(step_log[doc['link_to_step']])
        df_sheet_names = doc["file_details"]["sheet_names"]
        df_columns = doc["file_details"]["columns"]
        for sheet in df_sheet_names:
            book_sheet = book.worksheets[book.sheetnames.index(sheet)]  # getting the index
            rows_generator = book_sheet.values
            header_row = next(rows_generator)
            data_rows = [row for (_, row) in zip(range(5 - 1), rows_generator)]  # first 4 rows
            data_parse = pd.DataFrame(data_rows, columns=header_row)
            data_transpose = data_parse.transpose()
            data_reset = data_transpose.reset_index()
            data_column = data_reset.iloc[:, :len(doc["file_details"]["columns"])]
            data_drop = data_column.dropna()
            data_drop.columns = df_columns
            data_copy = data_drop.copy()
            data_copy.loc[:, "sub_category"] = sheet
            data_copy.reset_index(drop=True, inplace=True)
            list_df.append(data_copy)
        df = pd.concat(list_df, ignore_index=True)

        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_staples_attach_rate_data(doc):
    # developed for ATE-31
    # some new test comment
    global df
    ind = 0
    try:
        data = pd.ExcelFile(step_log[doc['link_to_step']], engine='openpyxl')
        df = data.parse(step_log[doc["file_details"]["xl_tab_from_step"]])

        recs = df[:doc["file_details"]["rows_to_keep"]].astype(str).drop(doc["file_details"]["drop_columns"],
                                                                         axis=1).to_dict('records')
        for r in recs:
            r[doc["file_details"]["col_to_replace_nans"]] = str(ind) if r[doc["file_details"][
                "col_to_replace_nans"]] == 'nan' else r['week of:']
            ind += 1
        df = pd.DataFrame(recs).set_index(doc["file_details"]["col_to_replace_nans"]).T

        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_walmart_sell_thru_data(doc):
    # DAN-1457 #
    global df
    try:
        data = pd.read_excel(step_log[doc['link_to_step']], sheet_name=0)

        ink_index = list(data["PRINTERS"]).index("INK")

        data = data[:ink_index].dropna(subset="PRINTERS") \
            .drop(columns=["WOS", "L4Wk Avg Sales", "Inventory", "In Stock %"]) \
            .rename(columns={"PRINTERS": "product_id", "Unnamed: 1": "product_desc"}) \
            .melt(id_vars=["product_id", "product_desc"], var_name="weekend_date", value_name="quantity") \
            .fillna(0)[["weekend_date", "product_id", "product_desc", "quantity"]]
        df = data
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())


def dataframe_load_office_depot_data(doc):
    '''{
		"step_id": "dataframe_load",
		"skip_rows": "1",
		"step_action": "dataframe_load_office_depot_data",
		"file_details": {
			"type": "xlsx",
			"col_use": "I-S",
			"sheet_names": [
				"Inkjet Stores Only",
				"Inkjet Direct Only",
				"Inkjet Direct BOPIS Only",
				"Laser Stores Only",
				"Laser Direct Only",
				"Laser Direct BOPIS Only"
			]
		},
		"link_to_step": "loading_file"
	},'''
    global df
    import re
    data = []
    try:
        skip_rows = int(doc["skip_rows"]) if "skip_rows" in doc.keys() and doc["skip_rows"] is not None else None

        for sheet_name in doc["file_details"]["sheet_names"]:

            sheet_name_split = sheet_name.split(" ")
            sheet_name_split.pop()
            sub_category = sheet_name_split[0]
            if sheet_name_split[-1] == "Stores":
                channel = "Store"
            else:
                channel = sheet_name_split[-1]

            # rows that required from final dataset
            if channel == "BOPIS":
                rows_required_column = "Total Retail Office Depot and OfficeMax Direct Buy Online Pickup In Store Only"
            else:
                rows_required_column = f"Total Retail Office Depot and OfficeMax {channel} Only"

            col_use = doc["file_details"]["col_use"].split("-")

            # Sheet columns alphabet from which alphabet to which eg: I-S
            use_cols = ",".join([chr(i) for i in range(ord(col_use[0]), ord(col_use[1]) + 1)])
            excel_data = pd.read_excel(step_log[doc['link_to_step']], sheet_name=sheet_name, skiprows=skip_rows,
                                       usecols=use_cols)
            excel_data_weekend = pd.read_excel(step_log[doc['link_to_step']], sheet_name=sheet_name, skiprows=skip_rows)

            # getting rows to remove based on the regex eg: 2 Jan 2021
            remove_rows = [i for i, c in enumerate(excel_data_weekend[rows_required_column]) if
                           not re.match(r"\d{2} \w{3,4} \d{4}", c)]
            # getting rows names
            wanted_row_names = [re.match(r"\d{2} \w{3,4} \d{4}", c).group() for i, c in
                                enumerate(excel_data_weekend[rows_required_column]) if
                                re.match(r"\d{2} \w{3,4} \d{4}", c)]
            # removing rows from dataframe
            excel_data.drop(remove_rows, inplace=True)

            # adding Columns
            excel_data.loc[:, 'sub_category'] = sub_category
            excel_data['weekending'] = wanted_row_names
            excel_data.loc[:, 'channel'] = channel
            columns_rename = {}
            replace_columns = list(
                map(lambda x: re.sub(r"[^A-Za-z]+", " ", x.lower().replace('.1', '')).strip().replace(" ", "_"),
                    excel_data.columns.values))
            for i, c in enumerate(excel_data.columns.values):
                columns_rename[c] = replace_columns[i]
            excel_data.rename(columns=columns_rename, inplace=True)
            data.append(excel_data)

        df = pd.concat(data, ignore_index=True)

        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_quill_weekly_report(doc):
    """Encoding attribute addition for a csv file and can be extended for other file type ATE-1871"""
    global df
    try:
        skip_rows = int(doc["skip_rows"]) if "skip_rows" in doc.keys() and doc["skip_rows"] is not None else None
        enc = doc['file_details']["encoding"] if "encoding" in doc['file_details'].keys() and doc['file_details'][
            'encoding'] is not None else None
        if doc['file_details']['type'] == 'csv':
            if 'quote_char' in doc['file_details']:
                if doc['file_details']['quote_char'] is not None:
                    df = pd.read_csv(step_log[doc['link_to_step']], sep=doc['file_details']['separator'],
                                     quotechar=doc['file_details']['quote_char'], skiprows=skip_rows, encoding=enc)
            else:
                df = pd.read_csv(step_log[doc['link_to_step']], sep=doc['file_details']['separator'],
                                 skiprows=skip_rows, encoding=enc)
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_turkey_tekra_seller(doc):
    """Getting the latest sheet in the excel file ATE-1635"""
    global df
    try:
        skip_rows = int(doc["skip_rows"]) if "skip_rows" in doc.keys() and doc["skip_rows"] is not None else None
        df = pd.ExcelFile(step_log[doc['link_to_step']], engine='openpyxl')
        sheet_name = df.sheet_names[-1]
        df = df.parse(sheet_name, skiprows=skip_rows)
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_staples_weekly_subcategory_performance(doc):
    import numpy as np
    global df
    try:
        sheet_name = doc["file_details"]["sheet_name"]  # "YOYHPSUMMARY"
        # Dataset Load
        dataset = pd.read_excel(step_log[doc['link_to_step']], sheet_name=sheet_name,
                                header=None)
        # Dataset extract subcategory
        dataset_subcategory = pd.read_excel(step_log[doc['link_to_step']],
                                            sheet_name=sheet_name,
                                            skiprows=1, usecols=list(range(3, len(dataset.columns)))).to_dict("list")
        # Dataset extract date
        date_extract_dataset = pd.read_excel(step_log[doc['link_to_step']],
                                             sheet_name=sheet_name,
                                             header=None, skiprows=2)

        sub_category_row = dataset.iloc[0, :].values
        sub_category_unpatterned = []
        for i in range(len(sub_category_row)):
            if (type(sub_category_row[i]) == np.float64 or type(sub_category_row[i]) == float) and (
                    pd.isna(sub_category_row[i]) or np.isnan(sub_category_row[i])):
                pass
            else:
                if type(sub_category_row[i]) != type(pd.NaT):
                    sub_category_unpatterned.append(sub_category_row[i])

        sub_category = []
        units_sub_category_index = []
        cogs_sub_category_index = []
        for i in range(len(sub_category_unpatterned)):
            # print(i)
            sub_category_unpatterned[i] = sub_category_unpatterned[i].replace("–", "-")
            if len(sub_category_unpatterned[i].split("-")) >= 3:
                s = sub_category_unpatterned[i].split("-")[2].split(" ")[:-1]
                if "units" in s[-1].casefold():
                    units_sub_category_index.append(i)
                if "cogs" in s[-1].casefold():
                    cogs_sub_category_index.append(i)
                s.insert(0, sub_category_unpatterned[i].split("-")[1].strip())
                # print(" ".join(s))
                sub_category.append(" ".join(s))
            else:
                if "units" in sub_category_unpatterned[i].split("-")[1].split(" ")[-1].casefold():
                    units_sub_category_index.append(i)
                if "cogs" in sub_category_unpatterned[i].split("-")[1].split(" ")[-1].casefold():
                    cogs_sub_category_index.append(i)
                sub_category.append(" ".join(sub_category_unpatterned[i].split("-")[1].split(" ")[1:-1]))

        sub_category_data = ["Contract", "Quill", ".com", "NAD", "Retail", "Total", "YOY Var %"]

        dataset_subcategory_contract_cogs = [
            dataset_subcategory[sub_category_data[0] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[0]] for i in range(0, 24, 2)]

        dataset_subcategory_contract_units = [
            dataset_subcategory[sub_category_data[0] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[0]] for i in range(1, 24, 2)]
        # print(dataset_subcategory_contract)
        # print(len(dataset_subcategory_contract))
        dataset_subcategory_quill_cogs = [
            dataset_subcategory[sub_category_data[1] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[1]] for i in range(0, 24, 2)]

        dataset_subcategory_quill_units = [
            dataset_subcategory[sub_category_data[1] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[1]] for i in range(1, 24, 2)]

        # print(dataset_subcategory_quill)
        # print(len(dataset_subcategory_quill))
        dataset_subcategory_usdotcom_cogs = [dataset_subcategory["US Dotcom"]] + [
            dataset_subcategory[sub_category_data[2] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[2]] for i in range(1, 23, 2)]

        dataset_subcategory_usdotcom_units = [
            dataset_subcategory[sub_category_data[2] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[2]] for i in range(0, 23, 2)]

        # print(dataset_subcategory_usdotcom)
        # print(len(dataset_subcategory_usdotcom))
        dataset_subcategory_nad_cogs = [
            dataset_subcategory[sub_category_data[3] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[3]] for i in range(0, 24, 2)]

        dataset_subcategory_nad_units = [
            dataset_subcategory[sub_category_data[3] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[3]] for i in range(1, 24, 2)]

        # print(dataset_subcategory_nad)
        # print(len(dataset_subcategory_nad))
        dataset_subcategory_usretail_cogs = [dataset_subcategory["US Retail"]] + [
            dataset_subcategory[sub_category_data[4] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[4]] for i in range(1, 23, 2)]

        dataset_subcategory_usretail_units = [
            dataset_subcategory[sub_category_data[4] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[4]] for i in range(0, 23, 2)]
        # print(dataset_subcategory_usretail)
        # print(len(dataset_subcategory_usretail))
        dataset_subcategory_total_cogs = [
            dataset_subcategory[sub_category_data[5] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[5]] for i in range(0, 24, 2)]

        dataset_subcategory_total_units = [
            dataset_subcategory[sub_category_data[5] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[5]] for i in range(1, 24, 2)]
        # print(dataset_subcategory_total)
        # print(len(dataset_subcategory_total))
        dataset_subcategory_yoy_cogs = [
            dataset_subcategory[sub_category_data[6] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[6]] for i in range(0, 24, 2)]

        dataset_subcategory_yoy_units = [
            dataset_subcategory[sub_category_data[6] + "." + str(i)] if i != 0 else dataset_subcategory[
                sub_category_data[6]] for i in range(1, 24, 2)]

        df_temp = []
        for i in range(len(sub_category) // 2):
            temp_df = pd.DataFrame(
                [dataset_subcategory_contract_cogs[i], dataset_subcategory_quill_cogs[i],
                 dataset_subcategory_usdotcom_cogs[i],
                 dataset_subcategory_nad_cogs[i], dataset_subcategory_usretail_cogs[i],
                 dataset_subcategory_total_cogs[i],
                 dataset_subcategory_yoy_cogs[i], dataset_subcategory_contract_units[i],
                 dataset_subcategory_quill_units[i],
                 dataset_subcategory_usdotcom_units[i], dataset_subcategory_nad_units[i],
                 dataset_subcategory_usretail_units[i],
                 dataset_subcategory_total_units[i], dataset_subcategory_yoy_units[i]]).T
            temp_df.rename(
                columns={0: "cogs_contract", 1: "cogs_quill", 2: "cogs_online", 3: "cogs_nad", 4: "cogs_store",
                         5: "cogs_total", 6: "cogs_yoy",
                         7: "units_contract", 8: "units_quill", 9: "units_online", 10: "units_nad",
                         11: "units_store", 12: "units_total", 13: "units_yoy"},
                inplace=True)
            temp_df.insert(0, "sub_category", sub_category[i])
            temp_df.insert(0, 'day', date_extract_dataset.iloc[:, [2]].values)

            df_temp.append(temp_df)

        df = pd.concat(df_temp)

    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_staples_weekly_subcategory(doc):
    global df
    try:
        skip_rows = int(doc["skip_rows"]) if "skip_rows" in doc.keys() and doc["skip_rows"] is not None else None
        sheet_name = doc["file_details"]["sheet_name"]
        col_use = doc["file_details"]["col_use"].split("-")
        category_col = [(lambda x: x + "_")(i) for i in doc["file_details"]["category_column_names"]]

        # Sheet columns alphabet from which alphabet to which eg: I-S
        use_cols = ",".join([chr(i) for i in range(ord(col_use[0]), ord(col_use[1]) + 1)])
        excel_data = pd.read_excel(step_log[doc['link_to_step']], sheet_name=sheet_name, skiprows=skip_rows,
                                   usecols=use_cols)
        default_col_name = ["CONTRACT", "QUILL", "ONLINE", "NAD", "STORE", "TOTAL", "YOY"]
        new_col_names = []
        for i in category_col:
            for j in default_col_name:
                new_col_names.append(i + j)
        excel_data.rename(columns={"Unnamed: 2": "day"}, inplace=True)
        excel_data.rename(columns=dict(zip(excel_data.columns[1:], new_col_names)), inplace=True)
        excel_data.columns = excel_data.columns.str.lower()
        df = excel_data.head(150 - (skip_rows + 1))
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_load_file_partitioned_data(doc):
    ''' Please use this function only if the data has two different tables in one single sheet and
    use skip_rows to differentiating and this function is build for ticket ATE-784
    Eg:{
		"step_id": "dataframe_load",
		"skip_rows": "11",
		"step_action": "dataframe_load_file_partitioned_data",
		"file_details": {
			"type": "xls",
			"nrows": "35",
			"sheet_name": "December",
			"column_names": "Product_ID,Item,Product_Page_Views,Unique_Visitor,Orders,Conversion_by_Unique_Visitor",
			"email_details": "subject"
		},
		"link_to_step": "loading_file"
	}
	'''
    import math
    global df
    link_to_step = [x for x in step_log_status if x['step_id'] == doc['link_to_step']][0]

    skip_rows = int(doc["skip_rows"]) if "skip_rows" in doc.keys() and doc["skip_rows"] is not None else None

    if link_to_step['status'] != 'OK':
        pass
    else:
        try:
            col_names = doc["file_details"].get("column_names").split(",")
            if doc['file_details'].get('email_details'):
                email_details = step_log['email_details'][doc['file_details'].get('email_details')]
                sheet_name = email_details.split("-")[-1].strip().split(" ")[0]
                email_details_subject_key = email_details.split("-")[-1].strip()[:3] + "-" + email_details.split("-")[
                                                                                                 -1].strip()[-4:]

            if doc['file_details']['type'] == 'csv':
                pass
            elif doc['file_details']['type'] == 'xls':
                # Read all sheets in your File and find sheet name
                df = pd.read_excel(step_log[doc['link_to_step']], sheet_name=sheet_name, skiprows=skip_rows,
                                   names=col_names)

            # convert dataframe to a record rows of data
            record_data = df.to_dict("records")
            row = 0
            data = []

            # while loop because we don't know when the data is null
            if col_names:
                while True:
                    if math.isnan(record_data[row][col_names[0]]):
                        break
                    data.append(record_data[row])
                    row += 1
            # else:
            #     col_n = list(record_data[0].keys())[0]
            #     while True:
            #         if math.isnan(record_data[row][col_n]):
            #             break
            #         data.append(record_data[row])
            #         row += 1

            df = pd.DataFrame(data)
            if doc['file_details'].get('email_details'):
                df['date'] = email_details_subject_key

            status_update(doc)
            return True
        except Exception:
            status_update(doc, traceback.format_exc().splitlines())
            return False


def dataframe_load_downloaded_file(doc):
    global df
    # if step_log_status[doc['link_to_step']]['status'] != 'OK':
    link_to_step = [x for x in step_log_status if x['step_id'] == doc['link_to_step']][0]

    skip_rows = int(doc["skip_rows"]) if "skip_rows" in doc.keys() and doc["skip_rows"] is not None else None

    if link_to_step['status'] != 'OK':
        pass
    else:
        try:
            if doc['file_details']['type'] == 'csv':
                if 'quote_char' in doc['file_details']:
                    if doc['file_details']['quote_char'] is not None:
                        df = pd.read_csv(step_log[doc['link_to_step']], sep=doc['file_details']['separator'],
                                         quotechar=doc['file_details']['quote_char'], skiprows=skip_rows,
                                         encoding_errors='replace')
                else:
                    df = pd.read_csv(step_log[doc['link_to_step']], sep=doc['file_details']['separator'],
                                     skiprows=skip_rows, encoding_errors='replace')
                    # step_log[doc['step_id']]
            elif doc['file_details']['type'] == 'xls':
                # Read all sheets in your File and find sheet name
                df = pd.read_excel(step_log[doc['link_to_step']], sheet_name=doc['file_details']['sheet_name'],
                                   skiprows=skip_rows)
            elif doc['file_details']['type'] == 'xlsx':
                df = pd.ExcelFile(step_log[doc['link_to_step']], engine='openpyxl')
                # df = df.parse(doc['file_details']['sheet_name'])
                # find sheet name
                if len(df.sheet_names) > 1:
                    sheet_name = \
                        [pg for pg in df.sheet_names if (doc['file_details']['sheet_name']).lower() in pg.lower()][0]
                else:
                    sheet_name = df.sheet_names[0]
                df = df.parse(sheet_name, skiprows=skip_rows)

            status_update(doc)
            return True
        except Exception:
            status_update(doc, traceback.format_exc().splitlines())
            return False


def dataframe_slicing_start_end_row(doc):
    # after parsing to dataframe removing the rows that are before and after the actual data
    global df
    starting_index = int(doc["starting_index"]) if "starting_index" in doc.keys() and doc[
        "starting_index"] is not None else 0
    ending_index = int(doc["starting_index"]) if "starting_index" in doc.keys() and doc[
        "starting_index"] is not None else len(df)
    try:
        for index, row in enumerate(df.iloc[:, 0]):
            if pd.isna(row):
                ending_index = index  # ending index where there is nan data for case of the dataframe
                break
        df = df.iloc[starting_index:ending_index, :]

        status_update(doc)
        return True

    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def file_remove_downloaded(doc):
    # This function will
    global tmp_downloads

    if len(os.listdir(tmp_downloads)) > 0:
        # os.remove(tmp_downloads + path_separator + os.listdir(tmp_downloads)[-1])
        # shutil.rmtree(tmp_downloads)
        for filename in os.listdir(tmp_downloads):
            os.remove(os.path.join(tmp_downloads, filename))
        status_update(doc)
        step_log[doc['step_id']] = None
        return True
    else:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def file_send_by_email(doc):
    pass


def file_remove_file(doc):
    global tmp_downloads
    if doc['file_name'] in os.listdir(tmp_downloads):
        os.remove(tmp_downloads + path_separator + doc['file_name'])
        status_update(doc)
        step_log[doc['step_id']] = None
        return True
    else:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_add_column(doc):
    global df
    val = None
    try:
        # val = doc['column']['value']['static'] if 'static' in doc['column']['value'] and doc['column']['value']['static'] is not None else step_log[doc['column']['value']['step']]
        if 'static' in doc['column']['value'] and doc['column']['value']['static'] is not None:
            val = doc['column']['value']['static']
        elif type(doc['column']['value']['step']) == list:
            for st in doc['column']['value']['step']:
                try:
                    f_val = str(re.sub(r'\W', '_', str(step_log[st]).lower()))
                except:
                    continue
                if val is None:
                    val = f_val
                else:
                    val += '_' + f_val
        else:
            val = step_log[doc['column']['value']['step']]

        if 'find_in_step_dict' in doc['column'] and doc['column']['find_in_step_dict'] is not None:
            val = step_log[doc['column']['find_in_step_dict']][val]

        df.insert(
            loc=doc['column']['index'],
            column=doc['column']['name'],
            value=val
        )
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_drop_column(doc):
    global df
    try:
        if doc['match_wildcard']:
            # find all columns in df like
            doc['value'] = [col for col in df.columns if doc['value'].lower() in col.lower()]
        df.drop(columns=doc['value'], inplace=True)
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_save(doc):
    global df, tmp_downloads
    try:
        f_name = None

        if type(doc['file_details']['name']['value']) == list:
            for st in doc['file_details']['name']['value']:
                try:
                    f_val = str(re.sub(r'\W', '_', str(step_log[st]).lower()))
                except:
                    continue
                if f_name is None:
                    f_name = f_val
                else:
                    f_name += '_' + f_val
        else:
            f_name = doc['file_details']['name']['value']

        file_name_and_path = (tmp_downloads
                              + path_separator
                              + f_name
                              + (step_log[doc['file_details']['name']['sufix_from_step']] if 'sufix_from_step' in
                                                                                             doc['file_details'][
                                                                                                 'name'].keys() and
                                                                                             doc['file_details'][
                                                                                                 'name'][
                                                                                                 'sufix_from_step'] is not None else '')
                              + '.'
                              + doc['file_details']['type']
                              )

        if doc['file_details']['type'] == 'csv':
            df.to_csv(
                file_name_and_path,
                index=False,
                sep=doc['file_details']['separator']
            )
        elif doc['file_details']['type'] == 'parquet':
            df.to_parquet(
                file_name_and_path,
                compression='gzip',
                index=False
            )
        step_log[doc['step_id']] = file_name_and_path
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_column_rename(doc):
    global df
    # Given a dict with mappings between current and desired column names this function will do a column rename
    # In case dict columns are not present an error will not be raised.
    try:
        if 'column_mapping' in doc.keys() and doc['column_mapping'] is not None:
            df.rename(columns=doc['column_mapping'], inplace=True)
            status_update(doc)
            return True
        else:
            status_update(doc)
            return False
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_slice_by_column(doc):
    global df
    # The system will return a df of X rows. X is determined by the number of records in the given column_name var.

    try:
        df = df[:df[doc['column_name']].count()]
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def dataframe_slice_columns(doc):
    global df
    # The system will return a df of column names mentioned in doc and on the same order mentioned in doc

    try:
        df = df[doc['column_names']]
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def step_execution(step):
    if step['step_action'] == 'sleep':
        sleep(step)
    elif step['step_action'] == 'file_get_downloaded':
        file_get_downloaded(step)
    elif step['step_action'] == 'file_remove_downloaded':
        file_remove_downloaded(step)
    elif step['step_action'] == 'get_date':
        get_date(step)
    elif step['step_action'] == 'file_remove_file':
        file_remove_file(step)
    elif step['step_action'] == 'dataframe_load_downloaded_file':
        dataframe_load_downloaded_file(step)
    elif step['step_action'] == 'dataframe_slicing_start_end_row':
        dataframe_slicing_start_end_row(step)
    elif step['step_action'] == 'dataframe_add_column':
        dataframe_add_column(step)
    elif step['step_action'] == 'dataframe_save':
        dataframe_save(step)
    elif step['step_action'] == 'set_variable':
        set_variable(step)
    elif step['step_action'] == 'variable_str_edit':
        variable_str_edit(step)
    elif step['step_action'] == 'text_if_then':
        text_if_then(step)
    elif step['step_action'] == 'step_process_stop':
        step_process_stop(step)
    elif step['step_action'] == 'file_push_to_s3':
        file_push_to_s3(step)
    elif step['step_action'] == 'dataframe_column_rename':
        dataframe_column_rename(step)
    elif step['step_action'] == 'dataframe_slice_by_column':
        dataframe_slice_by_column(step)
    elif step['step_action'] == 'dataframe_slice_columns':
        dataframe_slice_columns(step)
    elif step['step_action'] == 'cal_date_delta':
        cal_date_delta(step)
    # elif step['step_action'] == 'email_get_attachment':
    # email_get_attachment(step)
    # elif step['step_action'] == 'email_get_summary':
    # email_get_summary(step)
    elif step['step_action'] == 'dataframe_load_staples_segment':
        dataframe_load_staples_segment(step)
    elif step['step_action'] == 'dataframe_drop_column':
        dataframe_drop_column(step)
    elif step['step_action'] == 'dataframe_load_ink_toner_online_metrics':
        dataframe_load_ink_toner_online_metrics(step)
    elif step['step_action'] == 'dataframe_load_best_buy_monthly_data':
        dataframe_load_best_buy_monthly_data(step)
    elif step['step_action'] == 'dataframe_load_office_depot_data':
        dataframe_load_office_depot_data(step)
    elif step['step_action'] == 'dataframe_load_staples_weekly_subcategory':
        dataframe_load_staples_weekly_subcategory(step)
    elif step['step_action'] == 'dataframe_load_staples_weekly_subcategory_performance':
        dataframe_load_staples_weekly_subcategory_performance(step)
    elif step['step_action'] == 'dataframe_load_file_partitioned_data':
        dataframe_load_file_partitioned_data(step)
    elif step['step_action'] == 'process_loop_thru':
        process_loop_thru(step)
    elif step['step_action'] == 'dataframe_load_walmart_sell_thru_data':
        dataframe_load_walmart_sell_thru_data(step)
    # elif step['step_action'] == 'dataframe_load_best_buy_pdp_data':
    # dataframe_load_best_buy_pdp_data(step)
    elif step['step_action'] == 'dataframe_load_staples_attach_rate_data':
        dataframe_load_staples_attach_rate_data(step)
    elif step['step_action'] == 'dataframe_load_turkey_tekra_seller':
        dataframe_load_turkey_tekra_seller(step)
    elif step['step_action'] == 'dataframe_load_staples_attach_reporting':
        dataframe_load_staples_attach_reporting(step)
    elif step['step_action'] == 'unzip_zipping_file':
        unzip_zipping_file(step)
    elif step['step_action'] == 'dataframe_load_quill_weekly_report':
        dataframe_load_quill_weekly_report(step)
    elif step['step_action'] == 'dataframe_load_sams_club_revenue':
        dataframe_load_sams_club_revenue(step)
    elif step['step_action'] == 'dataframe_insert_file_name':
        dataframe_insert_file_name(step)


def sleep(doc):
    # this function will pause the execution for a given period of time.
    time.sleep(doc['seconds'])
    status_update(doc)


def status_update(doc, error=None):
    nr = {}
    nr['ts'] = str(datetime.now())

    if error is None:
        print('Step {} completed'.format(doc['step_id']))
        nr['status'] = 'OK'
    else:
        print('Step {} falied'.format(doc['step_id']))
        nr['status'] = 'ERROR'
        nr['trace'] = error

    nr['step_id'] = doc['step_id']
    step_log_status.append(nr)


def process_loop_thru(doc):
    try:
        for el in step_log[doc['link_to_step']]:
            step_log[doc['step_id']] = el
            step_loop_thru(doc['loop_steps'])
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def step_loop_thru(doc):
    for el in doc:
        if stop_process is False:
            step_execution(el)


def step_process_stop(doc):
    global stop_process

    stop_process = True
    status_update(doc, doc['reason'])
    return True


def text_if_then(doc):
    try:
        condition_check = False
        if doc['condition']['validation_type'] == 'eq':
            condition_check = True if step_log[doc['link_to_step']] == (
                doc['condition']['validation_right_side']['static'] if doc['condition']['validation_right_side'][
                                                                           'step'] is None else step_log[
                    doc['condition']['validation_right_side']['step']]) else False
        elif doc['condition']['validation_type'] == 'gt':
            condition_check = True if step_log[doc['link_to_step']] > (
                doc['condition']['validation_right_side']['static'] if doc['condition']['validation_right_side'][
                                                                           'step'] is None else step_log[
                    doc['condition']['validation_right_side']['step']]) else False
        elif doc['condition']['validation_type'] == 'lt':
            condition_check = True if step_log[doc['link_to_step']] < (
                doc['condition']['validation_right_side']['static'] if doc['condition']['validation_right_side'][
                                                                           'step'] is None else step_log[
                    doc['condition']['validation_right_side']['step']]) else False
        elif doc['condition']['validation_type'] == 'like':
            condition_check = True if (doc['condition']['validation_right_side']['static'] if
                                       doc['condition']['validation_right_side']['step'] is None else step_log[
                doc['condition']['validation_right_side']['step']]) in step_log[doc['link_to_step']] else False
        elif doc['condition']['validation_type'] == 'len_eq':
            condition_check = True if len(step_log[doc['link_to_step']]) == (
                doc['condition']['validation_right_side']['static'] if doc['condition']['validation_right_side'][
                                                                           'step'] is None else step_log[
                    doc['condition']['validation_right_side']['step']]) else False
        elif doc['condition']['validation_type'] == 'len_gt':
            condition_check = True if len(step_log[doc['link_to_step']]) > (
                doc['condition']['validation_right_side']['static'] if doc['condition']['validation_right_side'][
                                                                           'step'] is None else step_log[
                    doc['condition']['validation_right_side']['step']]) else False

        if condition_check:
            step_loop_thru(doc['actions_if_true'])
        else:
            step_loop_thru(doc['actions_if_false'])
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def get_date(doc):
    try:
        step_log[doc['step_id']] = datetime.strftime(
            datetime.now() + timedelta(doc['delta'] if 'delta' in doc and doc['delta'] is not None else 0),
            doc['date_format'] if doc['date_format'] is not None else '%Y-%m-%d'
        )
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def set_variable(doc):
    try:
        step_log[doc['step_id']] = doc['value']
        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def variable_str_edit(doc):
    try:
        tmp_val = str(step_log[doc['link_to_step']])
        for key in doc['actions'].keys():

            if key == 'replace':
                if doc['actions'][key]['is_regex']:
                    tmp_val = re.sub(doc['actions'][key]['old_val'], doc['actions'][key]['new_val'], tmp_val)
                else:
                    tmp_val = tmp_val.replace(doc['actions'][key]['old_val'], doc['actions'][key]['new_val'])
            elif key == 'to_lower_char' and doc['actions'][key]:
                tmp_val = tmp_val.lower()
            elif key == 'to_upper_char' and doc['actions'][key]:
                tmp_val = tmp_val.upper()
            elif key == 'strip' and doc['actions'][key]:
                tmp_val = tmp_val.strip()
            elif key == 'len' and doc['actions'][key]:
                tmp_val = len(tmp_val)
            elif key == 'sub_string':
                tmp_val = tmp_val[doc['actions'][key]['start_index']:doc['actions'][key]['last_index']]
            elif key == 'split':
                tmp_val = tmp_val.split(doc['actions'][key]['split_by'])[doc['actions'][key]['return_index']]

        step_log[doc['step_id']] = tmp_val

        status_update(doc)
        return True
    except Exception:
        status_update(doc, traceback.format_exc().splitlines())
        return False


def send_email():
    global env
    SENDER = "pan.coe.cloud.ops@hp.com"
    RECIPIENT = ([str(i) for i in contact.split(",")])
    CHARSET = "UTF-8"
    SUBJECT = "Email Crawler Lambda Error"

    BODY_HTML = """<html>
            <head></head>
            <body>Hello,
            </br></br>There was an error during the crawler execution for scraper_id - {} , job_id - {}. 
            </br>Please check database for error details.
            </br></br> 
            </br></br>Best regards
            </body>
            </html>""".format(source_id, job_id)

    client = boto3.client('ses', region_name='us-west-2')
    # Try to send the email.
    try:
        # Provide the contents of the email.
        response = client.send_email(
            Destination={
                'ToAddresses': RECIPIENT
            },
            Message={
                'Body': {
                    'Html': {
                        'Charset': CHARSET,
                        'Data': BODY_HTML,
                    }
                },
                'Subject': {
                    'Charset': CHARSET,
                    'Data': SUBJECT,
                },
            },
            Source=SENDER,
        )
    # Display an error if something goes wrong.
    except Exception as e:
        print(e.response['Error']['Message'])
        # update logs
        rs = {}
        rs = {"update_task": 'update_log', "job_id": job_id, "source_id": source_id, "source": source,
              "log_type": "ERROR", "log_details": [{"status": "error", "msg": str(e.response['Error']['Message'])}]}
        db_query(rs)

    else:
        print("Email sent! Message ID:")
        print(response['MessageId'])
